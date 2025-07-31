import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font
from openpyxl.utils import get_column_letter
import os
import re

def letra_para_indice(coluna):
    coluna = coluna.upper()
    indice = 0
    for c in coluna:
        indice = indice * 26 + (ord(c) - ord('A') + 1)
    return indice - 1

def ajustar_largura_coluna(ws, col_idx):
    max_length = 0
    col_letter = get_column_letter(col_idx)
    for cell in ws[col_letter]:
        if cell.value:
            max_length = max(max_length, len(str(cell.value)))
    ws.column_dimensions[col_letter].width = max_length + 2

def ocultar_colunas(ws, start_col):
    max_col = ws.max_column
    for col in range(start_col, max_col + 1):
        ws.column_dimensions[get_column_letter(col)].hidden = True

def processar_planilha(caminho_arquivo, hora_ref=None):
    # Se hora_ref não for fornecida, utiliza a hora atual
    if hora_ref is None:
        hora_ref = datetime.now().hour

    # Gerar nome do arquivo de saída com timestamp
    timestamp = datetime.now().strftime("%Y-%m-%d_%H-%M")
    nome_base = os.path.splitext(os.path.basename(caminho_arquivo))[0]
    arquivo_saida = f"{nome_base}_{timestamp}.xlsx"

    # Mapeamento das colunas por dia do mês
    colunas_por_dia = {
        1: "H", 2: "K", 3: "L", 4: "P", 5: "R", 6: "S", 7: "T", 8: "V",
        9: "X", 10: "Z", 11: "AA", 12: "AB", 13: "AC", 14: "AE", 15: "AF",
        16: "AI", 17: "AK", 18: "AM", 19: "AN", 20: "AO", 21: "AP", 22: "AQ",
        23: "AR", 24: "AT", 25: "AU", 26: "AV", 27: "AW", 28: "AX", 29: "BA",
        30: "BC", 31: "BD"
    }

    agora = datetime.now()
    dia_atual = min(agora.day, 31)
    mes_atual = agora.month
    ano_atual = agora.year % 100

    def obter_indice_col(col_letra):
        return letra_para_indice(col_letra)

    col_idx_hoje = obter_indice_col(colunas_por_dia[dia_atual])
    todas_colunas_idx = [obter_indice_col(col) for col in colunas_por_dia.values()]

    # Lê as abas que começam com "Page"
    xls = pd.ExcelFile(caminho_arquivo)
    abas = [aba for aba in xls.sheet_names if aba.startswith("Page")]

    dados = []

    for aba in abas:
        try:
            df = pd.read_excel(xls, sheet_name=aba, header=None)

            equipamento_raw = str(df.iat[8, 13]) if pd.notna(df.iat[8, 13]) else ""
            equipamento = equipamento_raw[5:9] if len(equipamento_raw) >= 9 else ""

            faixa = df.iat[9, 13] if pd.notna(df.iat[9, 13]) else ""
            qtd_total = df.iat[39, col_idx_hoje] if pd.notna(df.iat[39, col_idx_hoje]) else 0

            ultima_descarga_val = None
            linha_ultima = None
            col_ultima = None
            dia_ultima_descarga = None

            for idx_col, col in reversed(list(enumerate(todas_colunas_idx))):
                col_values = df.iloc[15:39, col]
                valores_validos = col_values[(col_values.notna()) & (col_values > 0)]
                if not valores_validos.empty:
                    ultima_descarga_val = valores_validos.iloc[-1]
                    linha_ultima = valores_validos.index[-1]
                    col_ultima = col
                    dia_ultima_descarga = list(colunas_por_dia.keys())[idx_col]
                    break

            if linha_ultima is not None:
                horario_completo = str(df.iat[linha_ultima, 1])
                match = re.search(r'\b\d{2}:\d{2}\b', horario_completo)
                horario_formatado = match.group(0) if match else horario_completo
            else:
                horario_formatado = ""

            if dia_ultima_descarga and horario_formatado:
                try:
                    data_descarga = datetime.strptime(
                        f"{dia_ultima_descarga:02d}/{mes_atual:02d}/{ano_atual:02d} {horario_formatado}",
                        "%d/%m/%y %H:%M"
                    )
                    data_formatada = data_descarga.strftime("%d/%m/%y %H:%M")
                except:
                    data_descarga = None
                    data_formatada = ""
            else:
                data_descarga = None
                data_formatada = ""

            if ultima_descarga_val is None:
                status = "Sem descargas no mês"
            elif data_descarga and data_descarga.date() != agora.date():
                status = "Não descarregou"
            else:
                if data_descarga and data_descarga.hour >= hora_ref:
                    status = "Descarregou"
                else:
                    status = "Não descarregou"

            alerta = "Atenção" if int(qtd_total) < 600 else ""

            dados.append({
                "Equipamento": equipamento,
                "Faixa": faixa,
                "QTD_Total": int(qtd_total),
                "Última Descarga": ultima_descarga_val,
                "Último funcionamento": data_formatada,
                "Status": status,
                "Alerta": alerta
            })

        except Exception as e:
            print(f"Erro na aba {aba}: {e}")

    df_final = pd.DataFrame(dados)
    df_final.to_excel(arquivo_saida, index=False)

    wb = load_workbook(arquivo_saida)
    ws = wb.active
    ws.title = "ALERTA DE DESCARGA"

    ws.auto_filter.ref = ws.dimensions
    alinhamento_central = Alignment(horizontal='center', vertical='center')
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            cell.alignment = alinhamento_central
            if cell.value:
                max_length = max(max_length, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_length + 2

    vermelho = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    verde = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    vermelho_fonte = Font(color="9C0006")
    verde_fonte = Font(color="006100")

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
        status_cell = row[5]
        # Alerta está na próxima coluna (índice 6)
        if status_cell.value == "Descarregou":
            status_cell.fill = verde
            status_cell.font = verde_fonte
        elif status_cell.value == "Não descarregou":
            status_cell.fill = vermelho
            status_cell.font = vermelho_fonte
        elif status_cell.value == "Sem descargas no mês":
            for cell in row:
                cell.fill = vermelho
                cell.font = vermelho_fonte

    # Cria/atualiza aba RESUMO
    if "RESUMO" in wb.sheetnames:
        ws_resumo = wb["RESUMO"]
        wb.remove(ws_resumo)
    ws_resumo = wb.create_sheet("RESUMO")

    linha = 1

    from openpyxl.styles import Font
    titulo_font = Font(bold=True, underline="single", size=14)
    titulo_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    titulo1 = ws_resumo.cell(row=linha, column=1, value="Alerta de fluxo")
    titulo1.font = titulo_font
    titulo1.fill = titulo_fill
    linha += 1

    equipamentos_alerta = [d for d in dados if d["Alerta"] == "Atenção"]

    if equipamentos_alerta:
        for d in equipamentos_alerta:
            cell = ws_resumo.cell(row=linha, column=1)
            cell.value = f"{d['Equipamento']} / {d['Faixa']} - Está com {d['QTD_Total']} de fluxo"
            cell.alignment = Alignment(horizontal='left')
            cell.font = Font(bold=False)
            linha += 1
    else:
        cell = ws_resumo.cell(row=linha, column=1, value="Não há alertas de fluxo.")
        cell.alignment = Alignment(horizontal='left')
        linha += 1

    linha += 1

    titulo2 = ws_resumo.cell(row=linha, column=1, value="Sem descarregar")
    titulo2.font = titulo_font
    titulo2.fill = titulo_fill
    linha += 1

    equipamentos_sem_descarga = [d for d in dados if d["Status"] != "Descarregou"]

    if equipamentos_sem_descarga:
        for d in equipamentos_sem_descarga:
            cell = ws_resumo.cell(row=linha, column=1)
            if d["Status"] == "Sem descargas no mês":
                texto = f"{d['Equipamento']} / {d['Faixa']} - Sem descargas no mês"
            else:
                texto = f"{d['Equipamento']} / {d['Faixa']} - Último funcionamento: {d['Último funcionamento']}"
            cell.value = texto
            cell.alignment = Alignment(horizontal='left')
            cell.font = Font(bold=False)
            linha += 1
    else:
        cell = ws_resumo.cell(row=linha, column=1, value="Todos descarregaram corretamente.")
        cell.alignment = Alignment(horizontal='left')

    ajustar_largura_coluna(ws_resumo, 1)
    if ws_resumo.max_column > 1:
        ocultar_colunas(ws_resumo, 2)
   
    wb.save(arquivo_saida)
    print(f"\n✅ Relatório salvo como: {arquivo_saida}")
    return arquivo_saida
